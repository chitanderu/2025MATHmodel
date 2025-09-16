import sys
from pathlib import Path

def print_header(title: str):
    print("=" * 80)
    print(title)
    print("=" * 80)

def try_import_pandas():
    try:
        import pandas as pd  # type: ignore
        return pd
    except Exception:
        return None

def inspect_with_pandas(xlsx_path: Path):
    pd = try_import_pandas()
    if pd is None:
        return False
    try:
        xl = pd.ExcelFile(xlsx_path)
    except Exception:
        return False

    print_header(f"Workbook: {xlsx_path.name} — Sheets: {len(xl.sheet_names)}")
    for i, sheet in enumerate(xl.sheet_names):
        try:
            # Read a tiny sample to infer columns reliably and keep output small
            df = pd.read_excel(xlsx_path, sheet_name=sheet, nrows=5, engine=None)
            cols = [str(c) for c in df.columns.tolist()]
            print(f"[{i+1:02d}] Sheet: {sheet}")
            print(f"    Columns ({len(cols)}): {cols}")
            # Show up to 2 data rows to avoid verbose output
            if len(df) > 0:
                preview_rows = min(2, len(df))
                print("    Preview rows:")
                print(df.head(preview_rows).to_string(index=False))
        except Exception as e:
            print(f"[{i+1:02d}] Sheet: {sheet}")
            print(f"    (preview failed with pandas: {e})")
    return True

def inspect_with_openpyxl(xlsx_path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return False
    try:
        wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return False

    sheet_names = wb.sheetnames
    print_header(f"Workbook: {xlsx_path.name} — Sheets: {len(sheet_names)}")
    for i, name in enumerate(sheet_names):
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        # Assume first non-empty row as header; fallback to first row
        header = None
        for r in ws.iter_rows(min_row=1, max_row=min(5, max_row)):
            values = [cell.value for cell in r]
            if any(v is not None and str(v).strip() != "" for v in values):
                header = [str(v) if v is not None else "" for v in values]
                break
        if header is None:
            header = []
        print(f"[{i+1:02d}] Sheet: {name}")
        print(f"    Approx size: rows={max_row}, cols={max_col}")
        print(f"    Columns (from first non-empty row, {len(header)}): {header}")
        # Preview first two data rows after header
        header_row_idx = 1
        if header:
            # find header row index again
            for r_idx, r in enumerate(ws.iter_rows(min_row=1, max_row=min(5, max_row)), start=1):
                values = [cell.value for cell in r]
                if [str(v) if v is not None else "" for v in values] == header:
                    header_row_idx = r_idx
                    break
        data_start = header_row_idx + 1
        preview_rows_printed = 0
        for r in ws.iter_rows(min_row=data_start, max_row=min(data_start + 1, max_row)):
            values = [cell.value for cell in r]
            print(f"    Row: {[v for v in values]}")
            preview_rows_printed += 1
        if preview_rows_printed == 0:
            print("    (no data rows)")
    return True

def main():
    # Allow passing an explicit path; otherwise try common locations
    candidates = []
    if len(sys.argv) >= 2:
        candidates.append(Path(sys.argv[1]))
    # Prefer workbook placed under scripts/ (as per your IDE view)
    candidates.append(Path("scripts/数据_区域双碳目标与路径规划研究（含拆分数据表）.xlsx"))
    # Also try repo root
    candidates.append(Path("数据_区域双碳目标与路径规划研究（含拆分数据表）.xlsx"))

    path = None
    for p in candidates:
        if p.exists():
            path = p
            break

    if path is None:
        tried = " | ".join(str(p) for p in candidates)
        print("Could not find the Excel workbook. Tried: ", tried)
        sys.exit(1)

    # Prefer pandas for nicer previews; fallback to openpyxl
    if inspect_with_pandas(path):
        return
    if inspect_with_openpyxl(path):
        return
    print("Neither pandas nor openpyxl is available to inspect the workbook.")
    sys.exit(2)

if __name__ == "__main__":
    main()
